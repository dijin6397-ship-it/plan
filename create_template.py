import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "排程数据"

header_font = Font(bold=True, size=12)
header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

headers = ["检修阶段", "SBOP/工序", "工单名称", "作业时长(小时)"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, size=12, color="FFFFFF")
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

sample_data = [
    ["第一阶段检修", "工序A", "工单A-1", 2],
    ["第一阶段检修", "工序A", "工单A-2", 3],
    ["第一阶段检修", "工序B", "工单B-1", 1.5],
    ["第一阶段检修", "工序B", "工单B-2", 2.5],
    ["第二阶段检修", "工序C", "工单C-1", 4],
    ["第二阶段检修", "工序C", "工单C-2", 2],
    ["第二阶段检修", "工序D", "工单D-1", 3],
]

for row_idx, row_data in enumerate(sample_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 15

wb.save('static/template.xlsx')
print("模板文件已生成: static/template.xlsx")
